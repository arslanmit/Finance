#!/usr/bin/env python3
"""Refresh the default S&P 500 workbook from Yahoo Finance monthly data."""

from __future__ import annotations

import json
import shutil
import time
import urllib.error
import urllib.request
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.comments import Comment

DEFAULT_WORKBOOK = Path("data/sp500_raw_data.xlsx")
DEFAULT_SHEET = "Sheet1"
DEFAULT_SYMBOL = "500.PA"
EARLIEST_REQUEST_DATE = "2010-01-01"
EARLIEST_REAL_MONTH = pd.Timestamp("2010-06-01")
OVERLAP_MONTHS = ("2024-08-01", "2024-10-01")
NOVEMBER_OPEN_MONTH = pd.Timestamp("2024-11-01")
EXPECTED_HEADERS = ["date", "open", "high", "low", "close", "volume"]
REQUEST_HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "Accept": "application/json,text/plain,*/*",
    "Accept-Language": "en-US,en;q=0.9",
}
YAHOO_HISTORY_URL = "https://finance.yahoo.com/quote/500.PA/history/"
AMUNDI_FACTSHEET_URL = (
    "https://www.amundietf.lu/en/individual/products/equity/"
    "amundi-sp-500-swap-ucits-etf-eur-acc/lu1681048804"
)
YAHOO_CHART_URL_TEMPLATE = (
    "https://query1.finance.yahoo.com/v8/finance/chart/{symbol}"
    "?interval=1mo&period1={period1}&period2={period2}"
    "&includePrePost=false&events=div%2Csplits"
)


class RefreshError(Exception):
    """Raised when the live workbook refresh fails."""


@dataclass(frozen=True)
class RefreshSummary:
    """Summary of a successful workbook refresh."""

    symbol: str
    row_count: int
    min_date: str
    max_date: str
    backup_path: str


def build_chart_url(symbol: str, period1: int, period2: int) -> str:
    return YAHOO_CHART_URL_TEMPLATE.format(symbol=symbol, period1=period1, period2=period2)


def fetch_monthly_source(symbol: str = DEFAULT_SYMBOL) -> pd.DataFrame:
    period1 = int(pd.Timestamp(EARLIEST_REQUEST_DATE).timestamp())
    period2 = int(time.time())
    url = build_chart_url(symbol, period1, period2)
    last_error: Exception | None = None

    for attempt in range(3):
        request = urllib.request.Request(url, headers=REQUEST_HEADERS)
        try:
            with urllib.request.urlopen(request, timeout=30) as response:
                payload = json.load(response)
            break
        except (urllib.error.URLError, TimeoutError, json.JSONDecodeError) as exc:
            last_error = exc
            if attempt == 2:
                raise RefreshError(f"Failed to fetch live data for {symbol}: {exc}") from exc
            time.sleep(1 + attempt)
    else:
        raise RefreshError(f"Failed to fetch live data for {symbol}: {last_error}")

    result = payload.get("chart", {}).get("result")
    if not result:
        error = payload.get("chart", {}).get("error")
        raise RefreshError(f"Yahoo Finance returned no chart data for {symbol}: {error}")

    chart = result[0]
    timestamps = chart.get("timestamp", [])
    quote = chart.get("indicators", {}).get("quote", [])
    if not timestamps or not quote:
        raise RefreshError(f"Yahoo Finance returned incomplete quote data for {symbol}.")

    frame = pd.DataFrame(
        {
            "date": pd.to_datetime(timestamps, unit="s", utc=True)
            .tz_convert("Europe/Paris")
            .tz_localize(None)
        }
    )
    for column, values in quote[0].items():
        frame[column] = values

    monthly = frame[
        frame["date"].dt.hour.eq(0)
        & frame["date"].dt.minute.eq(0)
        & frame["date"].dt.second.eq(0)
    ].copy()
    monthly["date"] = monthly["date"].dt.normalize()
    monthly = monthly[["date", "open", "high", "low", "close", "volume"]]
    monthly = monthly.drop_duplicates(subset=["date"]).sort_values("date").reset_index(drop=True)
    monthly = monthly[monthly["date"] >= EARLIEST_REAL_MONTH].copy()

    if monthly.empty:
        raise RefreshError(f"Yahoo Finance returned no monthly rows for {symbol}.")

    numeric_columns = ["open", "high", "low", "close"]
    monthly[numeric_columns] = monthly[numeric_columns].astype(float).round(2)
    monthly["volume"] = monthly["volume"].fillna(0).round(0).astype(int)

    return monthly


def load_existing_workbook_data(workbook_path: Path) -> pd.DataFrame:
    dataframe = pd.read_excel(workbook_path, sheet_name=DEFAULT_SHEET)
    dataframe.columns = [str(column).strip().lower() for column in dataframe.columns]
    dataframe["date"] = pd.to_datetime(dataframe["date"]).dt.normalize()
    return dataframe


def validate_source_contiguity(source: pd.DataFrame) -> None:
    if source["date"].duplicated().any():
        raise RefreshError("Live monthly source contains duplicate dates.")

    expected = pd.date_range(source["date"].min(), source["date"].max(), freq="MS")
    missing = expected.difference(source["date"])
    if not missing.empty:
        preview = ", ".join(day.date().isoformat() for day in missing[:5])
        raise RefreshError(f"Live monthly source is missing dates: {preview}.")


def validate_overlap(existing: pd.DataFrame, source: pd.DataFrame) -> None:
    merged = existing.merge(source, on="date", suffixes=("_existing", "_source"))
    if merged.empty:
        raise RefreshError("Could not compare the current workbook to the live source.")

    completed = merged[
        merged["date"].between(pd.Timestamp(OVERLAP_MONTHS[0]), pd.Timestamp(OVERLAP_MONTHS[1]))
    ].sort_values("date")
    if len(completed) != 3:
        raise RefreshError(
            "Expected three completed overlap months (2024-08 through 2024-10) for validation."
        )

    for column in ("open", "high", "low", "close"):
        delta = (completed[f"{column}_existing"] - completed[f"{column}_source"]).abs().max()
        if float(delta) > 0.02:
            raise RefreshError(
                f"Live source validation failed for {column}; max delta was {float(delta):.4f}."
            )

    november = merged[merged["date"] == NOVEMBER_OPEN_MONTH]
    if november.empty:
        raise RefreshError("Expected an overlap row for 2024-11-01.")

    open_delta = abs(float(november.iloc[0]["open_existing"]) - float(november.iloc[0]["open_source"]))
    if open_delta > 0.02:
        raise RefreshError(
            f"Live source validation failed for 2024-11 open; delta was {open_delta:.4f}."
        )


def create_backup(workbook_path: Path) -> Path:
    backup_dir = Path("tmp/spreadsheets")
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup_name = f"{workbook_path.stem}.backup.{time.strftime('%Y%m%d-%H%M%S')}{workbook_path.suffix}"
    backup_path = backup_dir / backup_name
    shutil.copy2(workbook_path, backup_path)
    return backup_path


def find_last_populated_row(worksheet) -> int:
    last_row = 1
    for row_index in range(2, worksheet.max_row + 1):
        if worksheet.cell(row=row_index, column=1).value is not None:
            last_row = row_index
    return last_row


def write_workbook_rows(workbook_path: Path, source: pd.DataFrame) -> None:
    workbook = load_workbook(workbook_path)
    if workbook.sheetnames != [DEFAULT_SHEET]:
        raise RefreshError(f"Unexpected workbook sheets: {workbook.sheetnames}")

    worksheet = workbook[DEFAULT_SHEET]
    headers = [worksheet.cell(row=1, column=index).value for index in range(1, 7)]
    if headers != EXPECTED_HEADERS:
        raise RefreshError(f"Unexpected workbook headers: {headers}")

    last_populated_row = find_last_populated_row(worksheet)
    blank_style_row = min(last_populated_row + 1, worksheet.max_row)
    data_styles = {column: copy(worksheet.cell(row=2, column=column)._style) for column in range(1, 7)}
    blank_styles = {
        column: copy(worksheet.cell(row=blank_style_row, column=column)._style) for column in range(1, 7)
    }

    descending = source.sort_values("date", ascending=False).reset_index(drop=True)
    output_rows: list[list[object]] = []
    for record in descending.itertuples(index=False):
        output_rows.append(
            [
                record.date.to_pydatetime(),
                float(record.open),
                float(record.high),
                float(record.low),
                float(record.close),
                int(record.volume),
            ]
        )

    max_target_row = max(worksheet.max_row, len(output_rows) + 1)
    for row_index in range(2, max_target_row + 1):
        if row_index - 2 < len(output_rows):
            values = output_rows[row_index - 2]
            styles = data_styles
        else:
            values = [None, None, None, None, None, None]
            styles = blank_styles

        for column_index, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell._style = copy(styles[column_index])
            cell.value = value

    worksheet["A1"].comment = Comment(
        "Refreshed from Yahoo Finance 500.PA monthly history and the Amundi LU1681048804 factsheet on "
        f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.\n"
        f"{YAHOO_HISTORY_URL}\n"
        f"{AMUNDI_FACTSHEET_URL}\n"
        "Earliest reliable live month for this series: 2010-06-01.",
        "Codex",
    )
    workbook.save(workbook_path)


def refresh_default_workbook(workbook_path: Path = DEFAULT_WORKBOOK) -> RefreshSummary:
    if workbook_path != DEFAULT_WORKBOOK:
        raise RefreshError("Refresh is only supported for the default sp500 workbook.")
    if not workbook_path.exists():
        raise RefreshError(f"Refresh target does not exist: {workbook_path}")

    existing = load_existing_workbook_data(workbook_path)
    source = fetch_monthly_source()
    validate_source_contiguity(source)
    validate_overlap(existing, source)
    backup_path = create_backup(workbook_path)
    write_workbook_rows(workbook_path, source)

    return RefreshSummary(
        symbol=DEFAULT_SYMBOL,
        row_count=len(source),
        min_date=source["date"].min().date().isoformat(),
        max_date=source["date"].max().date().isoformat(),
        backup_path=str(backup_path),
    )
