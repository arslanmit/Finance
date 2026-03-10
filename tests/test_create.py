from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from finance_cli.create import (
    build_generated_dataset_path,
    create_and_register_symbol_dataset,
    normalize_symbol_slug,
)
from finance_cli.errors import CreationError
from finance_cli.registry import load_registry


def monthly_frame() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "date": pd.to_datetime(["2024-01-01", "2024-02-01", "2024-03-01"]),
            "open": [10.0, 12.0, 11.0],
            "high": [11.0, 13.0, 12.0],
            "low": [9.0, 11.0, 10.0],
            "close": [10.5, 12.5, 11.5],
            "volume": [100, 110, 120],
        }
    )


def write_registry(path: Path) -> None:
    path.write_text('{"datasets": []}\n', encoding="utf-8")


def test_normalize_symbol_slug() -> None:
    assert normalize_symbol_slug("500.PA") == "500_pa"
    assert normalize_symbol_slug(" SPY ") == "spy"
    assert normalize_symbol_slug("BRK.B") == "brk_b"


def test_build_generated_dataset_path() -> None:
    assert build_generated_dataset_path("spy") == Path("data/generated/spy.xlsx")


def test_create_dataset_writes_workbook_and_registry(tmp_path: Path) -> None:
    config_path = tmp_path / "datasets.json"
    write_registry(config_path)
    datasets = load_registry(config_path=config_path)

    dataset = create_and_register_symbol_dataset(
        datasets,
        "SPY",
        config_path=config_path,
        fetcher=lambda symbol: monthly_frame(),
    )

    assert dataset.id == "spy"
    assert dataset.path == "data/generated/spy.xlsx"
    assert dataset.refresh is not None
    assert dataset.refresh.symbol == "SPY"

    workbook_path = tmp_path / dataset.path
    workbook = load_workbook(workbook_path)
    assert workbook.sheetnames == ["Sheet1"]

    sheet = workbook["Sheet1"]
    headers = [sheet.cell(row=1, column=index).value for index in range(1, 7)]
    assert headers == ["date", "open", "high", "low", "close", "volume"]
    assert sheet.cell(row=2, column=1).value.date().isoformat() == "2024-03-01"

    loaded_registry = load_registry(config_path=config_path)
    assert [item.id for item in loaded_registry] == ["spy"]


def test_create_dataset_rejects_duplicate_id(tmp_path: Path) -> None:
    config_path = tmp_path / "datasets.json"
    write_registry(config_path)
    datasets = load_registry(config_path=config_path)
    create_and_register_symbol_dataset(
        datasets,
        "SPY",
        config_path=config_path,
        fetcher=lambda symbol: monthly_frame(),
    )

    with pytest.raises(CreationError, match="already exists"):
        create_and_register_symbol_dataset(
            datasets,
            "SPY",
            config_path=config_path,
            fetcher=lambda symbol: monthly_frame(),
        )


def test_create_dataset_rejects_existing_target_file(tmp_path: Path) -> None:
    config_path = tmp_path / "datasets.json"
    write_registry(config_path)
    target = tmp_path / "data" / "generated" / "spy.xlsx"
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text("existing", encoding="utf-8")
    datasets = load_registry(config_path=config_path)

    with pytest.raises(CreationError, match="already exists"):
        create_and_register_symbol_dataset(
            datasets,
            "SPY",
            config_path=config_path,
            fetcher=lambda symbol: monthly_frame(),
        )


def test_create_dataset_rejects_empty_data(tmp_path: Path) -> None:
    config_path = tmp_path / "datasets.json"
    write_registry(config_path)
    datasets = load_registry(config_path=config_path)

    with pytest.raises(CreationError, match="no monthly data"):
        create_and_register_symbol_dataset(
            datasets,
            "SPY",
            config_path=config_path,
            fetcher=lambda symbol: monthly_frame().iloc[0:0],
        )


def test_create_dataset_rejects_invalid_symbol_slug(tmp_path: Path) -> None:
    config_path = tmp_path / "datasets.json"
    write_registry(config_path)
    datasets = load_registry(config_path=config_path)

    with pytest.raises(CreationError, match="letter or number"):
        create_and_register_symbol_dataset(
            datasets,
            "!!!",
            config_path=config_path,
            fetcher=lambda symbol: monthly_frame(),
        )


def test_create_dataset_wraps_fetch_errors(tmp_path: Path) -> None:
    config_path = tmp_path / "datasets.json"
    write_registry(config_path)
    datasets = load_registry(config_path=config_path)

    with pytest.raises(CreationError, match="network down"):
        create_and_register_symbol_dataset(
            datasets,
            "SPY",
            config_path=config_path,
            fetcher=lambda symbol: (_ for _ in ()).throw(RuntimeError("network down")),
        )
