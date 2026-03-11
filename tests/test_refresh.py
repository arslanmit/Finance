from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from finance_cli.errors import RefreshError
from finance_cli.models import DatasetConfig, RefreshMetadata, ResolvedSource
from finance_cli.refresh import validate_refreshable_source, write_workbook_rows


def test_validate_refresh_rejects_custom_file(tmp_path: Path) -> None:
    csv_path = tmp_path / "sample.csv"
    csv_path.write_text("date,open\n2024-01-01,10\n", encoding="utf-8")
    source = ResolvedSource(input_path=csv_path, sheet_name=None, dataset=None)

    with pytest.raises(RefreshError, match="registered datasets"):
        validate_refreshable_source(source)


def test_validate_refresh_requires_sheet(tmp_path: Path) -> None:
    workbook_path = tmp_path / "sample.xlsx"
    workbook_path.write_text("", encoding="utf-8")
    dataset = DatasetConfig(
        id="sample",
        label="Sample",
        path="sample.xlsx",
        sheet=None,
        refresh=RefreshMetadata(provider="yahoo", symbol="500.PA"),
        base_dir=tmp_path,
    )
    source = ResolvedSource(input_path=workbook_path, sheet_name=None, dataset=dataset)

    with pytest.raises(RefreshError, match="configured Excel sheet"):
        validate_refreshable_source(source)


def test_validate_refresh_rejects_unsupported_provider(tmp_path: Path) -> None:
    workbook_path = tmp_path / "sample.xlsx"
    workbook_path.write_text("", encoding="utf-8")
    dataset = DatasetConfig(
        id="sample",
        label="Sample",
        path="sample.xlsx",
        sheet="Sheet1",
        refresh=RefreshMetadata(provider="custom", symbol="500.PA"),
        base_dir=tmp_path,
    )
    source = ResolvedSource(input_path=workbook_path, sheet_name="Sheet1", dataset=dataset)

    with pytest.raises(RefreshError, match="unsupported refresh provider"):
        validate_refreshable_source(source)


def test_write_workbook_rows_preserves_symbol_first_column(tmp_path: Path) -> None:
    workbook_path = tmp_path / "spy.xlsx"
    with pd.ExcelWriter(workbook_path) as writer:
        pd.DataFrame(
            {
                "symbol": ["SPY"],
                "date": [pd.Timestamp("2024-01-01")],
                "open": [10.0],
                "high": [11.0],
                "low": [9.0],
                "close": [10.5],
                "volume": [100],
            }
        ).to_excel(writer, sheet_name="Sheet1", index=False)

    source = pd.DataFrame(
        {
            "date": pd.to_datetime(["2024-02-01", "2024-03-01"]),
            "open": [12.0, 11.0],
            "high": [13.0, 12.0],
            "low": [11.0, 10.0],
            "close": [12.5, 11.5],
            "volume": [110, 120],
        }
    )

    write_workbook_rows(workbook_path, source, "Sheet1", "SPY")

    workbook = load_workbook(workbook_path)
    sheet = workbook["Sheet1"]
    assert sheet.cell(row=1, column=1).value == "symbol"
    assert sheet.cell(row=2, column=1).value == "SPY"
    assert sheet.cell(row=2, column=2).value.date().isoformat() == "2024-03-01"


def test_write_workbook_rows_adds_symbol_first_column_when_missing(tmp_path: Path) -> None:
    workbook_path = tmp_path / "nvda.xlsx"
    with pd.ExcelWriter(workbook_path) as writer:
        pd.DataFrame(
            {
                "date": [pd.Timestamp("2024-01-01")],
                "open": [10.0],
                "high": [11.0],
                "low": [9.0],
                "close": [10.5],
                "volume": [100],
            }
        ).to_excel(writer, sheet_name="Sheet1", index=False)

    source = pd.DataFrame(
        {
            "date": pd.to_datetime(["2024-02-01", "2024-03-01"]),
            "open": [12.0, 11.0],
            "high": [13.0, 12.0],
            "low": [11.0, 10.0],
            "close": [12.5, 11.5],
            "volume": [110, 120],
        }
    )

    write_workbook_rows(workbook_path, source, "Sheet1", "NVDA")

    workbook = load_workbook(workbook_path)
    sheet = workbook["Sheet1"]
    headers = [sheet.cell(row=1, column=index).value for index in range(1, 8)]
    assert headers == ["symbol", "date", "open", "high", "low", "close", "volume"]
    assert sheet.cell(row=2, column=1).value == "NVDA"
    assert sheet.cell(row=2, column=2).value.date().isoformat() == "2024-03-01"
